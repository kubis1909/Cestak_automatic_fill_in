import openpyxl

SLOT_RADKY = list(range(29,69, 2))


def vyplneni_hlavicky(ws, hlavicka):
    ws["F2"] = hlavicka["jmeno"]
    ws["E3"] = hlavicka["bydliste"]
    ws["H11"] = hlavicka["typ_dopravy"]
 
def zapis_cestu(ws, radek, datum, odjezd, prijezd, km):
    ws[f"A{radek}"] = datum
    ws[f"C{radek}"] = odjezd
    ws[f"C{radek + 1}"] = prijezd
    ws[f"G{radek}"] = km
    ws[f"B{radek}"] = "odjezd"
    ws[f"B{radek + 1}"] = "prijezd"
    ws[f"E{radek}"] = "*AUV"
    ws[f"F{radek}"] = 3.50
    
    
    

def vytvor_cestak(sablona_prazdna, vystup_cesta, hlavicka, dny_cesty):
    if len(dny_cesty) > len(SLOT_RADKY):
        raise ValueError(f"Cesták má místo jen pro {len(SLOT_RADKY)} cest," f"zadal jsi jen {len(dny_cesty)}")

    wb = openpyxl.load_workbook(sablona_prazdna)
    ws = wb["List1"]

    vyplneni_hlavicky(ws, hlavicka)

    for datum, radek in zip(dny_cesty, SLOT_RADKY):
        zapis_cestu(ws, radek, datum, hlavicka["odjezd"], hlavicka["prijezd"],hlavicka["km"])

    wb.save(vystup_cesta)


if __name__ == "__main__":

    hlavicka = {
        "jmeno": "Jakub Petrik",
        "bydliste": "Hradec Králové",
        "typ_dopravy": "Škoda Octavia, 7H7 8622",
        "odjezd": "Hradec Králové",
        "prijezd": "Náchod",
        "km": 100,
    }

    dny = ["1.3.", "3.3.", "5.3.", "8.3.", "10.3.", "12.3."]

    vytvor_cestak("sablona_prazdna.xlsx", "test_vystup.xlsx", hlavicka, dny)

    print("Hotovo, cestak napsan.")